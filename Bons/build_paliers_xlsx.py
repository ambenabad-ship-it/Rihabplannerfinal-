# -*- coding: utf-8 -*-
"""
Read the canonical gift list from
  GIFTS/PRODUCTS_GIFTS  - Last Version (1).xlsx (sheet: Latest)
and write a clean paliers.xlsx the creator can upload through the app's
'Upload paliers' button.

Cleanup applied to each gift cell:
  * strip ";price" suffixes (everything after a ';')
  * collapse multiline cells like 'Triporteur;20 000\nSamsung 50' into
    'Triporteur + Samsung 50'
  * trim whitespace
Output columns match the parser: Palier, Threshold, Cadeau 1, Cadeau 2, Cadeau 3.
"""
import os, re
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

SRC = '/sessions/vibrant-tender-mayer/mnt/GIFTS/PRODUCTS_GIFTS  - Last Version (1).xlsx'
DST = '/sessions/vibrant-tender-mayer/mnt/Bons/paliers.xlsx'


def clean(cell):
    if cell is None: return ''
    s = str(cell)
    # split multi-gift cells
    parts = re.split(r'\r?\n', s)
    out = []
    for p in parts:
        p = p.strip()
        if not p: continue
        # strip ';price' suffix
        if ';' in p:
            p = p.split(';', 1)[0].strip()
        out.append(p)
    return ' + '.join(out)


def main():
    src = openpyxl.load_workbook(SRC, data_only=True)
    if 'Latest' not in src.sheetnames:
        raise SystemExit('Latest sheet missing in source xlsx')
    s = src['Latest']

    rows = list(s.iter_rows(values_only=True))
    # Row 0 = header, row 1 = "Engagement" sub-header, rows 2..15 = data.
    palier_n = 0
    paliers = []
    for r in rows[2:]:
        thr = r[0]
        if thr is None: continue
        try:
            thr = float(thr)
        except Exception:
            continue
        palier_n += 1
        paliers.append({
            'palier': palier_n,
            'threshold': int(thr),
            'p1': clean(r[1]),
            'p2': clean(r[2]),
            'p3': clean(r[3]),
        })

    print('Built', len(paliers), 'paliers')
    for p in paliers[:3]:
        print('  P{palier} · {threshold:,} MAD'.format(**p))
        print('     1:', p['p1'])
        print('     2:', p['p2'])
        print('     3:', p['p3'])
    print('  ...')

    # Build the destination workbook.
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Paliers'

    headers = ['Palier', 'Threshold', 'Cadeau 1', 'Cadeau 2', 'Cadeau 3']
    ws.append(headers)
    for c in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill('solid', fgColor='D54B33')
        cell.alignment = Alignment(horizontal='center', vertical='center')

    for p in paliers:
        ws.append([p['palier'], p['threshold'], p['p1'], p['p2'], p['p3']])

    # Column widths.
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 14
    ws.column_dimensions['C'].width = 50
    ws.column_dimensions['D'].width = 50
    ws.column_dimensions['E'].width = 60
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.alignment = Alignment(vertical='top', wrap_text=True)

    wb.save(DST)
    print('Wrote', DST)


if __name__ == '__main__':
    main()
